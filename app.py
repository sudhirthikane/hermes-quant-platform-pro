import streamlit as st
import pandas as pd
import json
import os
import io
import requests
import concurrent.futures
import warnings
from datetime import datetime
import plotly.graph_objects as go
import streamlit.components.v1 as components
import yfinance as yf
from engine import fetch_data, calculate_ict_indicators, LOG_FILE
from backtester import run_backtest

# Openpyxl for advanced Excel formatting
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Indicator library with fallback if pandas_ta is not directly available
# Explicit WMA and RSI functions for guaranteed accuracy
def calc_wma(series, length=21):
    """Calculates Weighted Moving Average (WMA) with exact mathematical weights [1..N]."""
    weights = np.arange(1, length + 1)
    return series.rolling(length).apply(lambda x: np.dot(x, weights) / weights.sum(), raw=True)

def calc_rsi(series, length=9):
    """Calculates Relative Strength Index (RSI)."""
    delta = series.diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=length).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=length).mean()
    rs = gain / loss.replace(0, np.nan)
    return 100 - (100 / (1 + rs))

try:
    import pandas_ta as ta
except ImportError:
    class FallbackTA:
        @staticmethod
        def rsi(close_series, length=9):
            return calc_rsi(close_series, length)
        
        @staticmethod
        def wma(close_series, length=21):
            return calc_wma(close_series, length)
    ta = FallbackTA

warnings.filterwarnings("ignore")

# -----------------------------------------------------------------------------
# CUSTOM SCANNER DATA ACQUISITION & ENGINE HELPERS
# -----------------------------------------------------------------------------
@st.cache_data(ttl=3600)
def get_indian_equities():
    """Fetches active equity tickers from the NSE India archives."""
    url = "https://nsearchives.nseindia.com/content/equities/EQUITY_L.csv"
    headers = {'User-Agent': 'Mozilla/5.0'}
    try:
        r = requests.get(url, headers=headers, timeout=10)
        df = pd.read_csv(io.BytesIO(r.content))
        return (df['SYMBOL'] + ".NS").tolist()
    except Exception as e:
        st.sidebar.error(f"Failed to fetch NSE tickers: {e}")
        return []

@st.cache_data(ttl=3600)
def get_us_equities(limit=1000):
    """Fetches top US stock tickers using the official SEC API."""
    url = "https://www.sec.gov/files/company_tickers.json"
    headers = {'User-Agent': 'MarketScannerUI (your@email.com)'} 
    try:
        r = requests.get(url, headers=headers, timeout=10)
        data = r.json()
        tickers = [item['ticker'] for item in data.values()]
        return [t.replace('-', '-') for t in tickers][:limit]
    except Exception as e:
        st.sidebar.error(f"Failed to fetch US tickers: {e}")
        return []

def get_commodities():
    """Returns major global commodity futures."""
    return [
        "GC=F", "SI=F", "PL=F", "PA=F", "HG=F",  # Metals
        "CL=F", "HO=F", "NG=F", "RB=F", "BZ=F",  # Energy
        "ZC=F", "ZW=F", "ZS=F", "ZM=F", "ZL=F",  # Grains
        "KC=F", "CT=F", "SB=F", "CC=F", "OJ=F",  # Softs
        "LE=F", "HE=F", "GF=F", "LBS=F"          # Meats & Lumber
    ]

def check_stock_conditions(ticker, market_name, timeframe="1H"):
    """Scans a single ticker for the selected timeframe strategy conditions: RSI(9) > WMA(RSI(9), 21)."""
    try:
        tf_clean = timeframe.lower().strip()
        
        if tf_clean == "1d":
            df = yf.download(ticker, period="1y", interval="1d", progress=False)
            if df.empty:
                return None
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(0)
        else:
            df_1h = yf.download(ticker, period="60d", interval="1h", progress=False)
            if df_1h.empty:
                return None
            if isinstance(df_1h.columns, pd.MultiIndex):
                df_1h.columns = df_1h.columns.get_level_values(0)
                
            if tf_clean == "1h":
                df = df_1h
            else:
                resample_freq = "3h" if tf_clean == "3h" else "4h"
                ohlc_dict = {'Open': 'first', 'High': 'max', 'Low': 'min', 'Close': 'last', 'Volume': 'sum'}
                agg_dict = {col: ohlc_dict[col] for col in df_1h.columns if col in ohlc_dict}
                try:
                    df = df_1h.resample(resample_freq, origin='start').agg(agg_dict).dropna()
                except TypeError:
                    df = df_1h.resample(resample_freq).agg(agg_dict).dropna()
        
        if df.empty or len(df) < 32:
            return None

        # Ensure 1D pandas Series for Close & Volume
        close_s = df['Close'].squeeze() if isinstance(df['Close'], pd.DataFrame) else df['Close']
        volume_s = df['Volume'].squeeze() if isinstance(df['Volume'], pd.DataFrame) else df['Volume']

        # Step 1: Calculate 9-period RSI
        df['RSI_9'] = calc_rsi(close_s, length=9)
        
        # Step 2: Calculate 21-period WMA ON RSI_9 (WMA of RSI)
        df['WMA_RSI_21'] = calc_wma(df['RSI_9'], length=21)
        
        df.dropna(subset=['RSI_9', 'WMA_RSI_21'], inplace=True)
        
        if len(df) < 2:
            return None

        curr_candle = df.iloc[-1]

        curr_close = float(curr_candle['Close'])
        curr_rsi = float(curr_candle['RSI_9'])
        curr_wma_rsi = float(curr_candle['WMA_RSI_21'])
        curr_vol = float(curr_candle['Volume'])

        # Strategy Conditions: RSI(9) > WMA(RSI(9), 21), Price > 70, Volume > 200k
        rsi_above_wma = curr_rsi > curr_wma_rsi
        price_above_70 = curr_close > 70
        volume_high = curr_vol > 200000

        if rsi_above_wma and price_above_70 and volume_high:
            return {
                'Ticker': ticker,
                'Market': market_name,
                'Timeframe': timeframe,
                'Time': df.index[-1].strftime('%Y-%m-%d %H:%M:%S'),
                'Close': round(curr_close, 2),
                'Volume': int(curr_vol),
                'RSI (9)': round(curr_rsi, 2),
                'WMA(RSI, 21)': round(curr_wma_rsi, 2)
            }
    except Exception:
        pass
    return None

def create_formatted_excel(df):
    """Generates a styled Excel workbook in memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Scanner Results"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    zebra_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='DDDDDD'), 
                         right=Side(style='thin', color='DDDDDD'), 
                         top=Side(style='thin', color='DDDDDD'), 
                         bottom=Side(style='thin', color='DDDDDD'))

    # Apply Header Styling
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Apply Row Styling
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8), start=2):
        for cell in row:
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = zebra_fill
            
            # Number formats
            if cell.column_letter in ['E', 'H']:
                cell.number_format = '#,##0.00'
            elif cell.column_letter == 'F':
                cell.number_format = '#,##0'
            elif cell.column_letter == 'G':
                cell.number_format = '0.00'

    # Auto-adjust column widths
    column_widths = {'A': 15, 'B': 18, 'C': 12, 'D': 20, 'E': 12, 'F': 15, 'G': 10, 'H': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.auto_filter.ref = ws.dimensions

    # Save to BytesIO buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

    # Auto-adjust column widths
    column_widths = {'A': 15, 'B': 18, 'C': 20, 'D': 12, 'E': 15, 'F': 10, 'G': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.auto_filter.ref = ws.dimensions

    # Save to BytesIO buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def add_print_button(report_title="Professional Equity Report"):
    from datetime import datetime
    current_time_str = datetime.now().strftime("%d-%b-%Y %I:%M %p")
    st.markdown(f"""
        <div class="print-only-header" style="display: none;">
            <div style="font-size: 22px; font-weight: bold; text-align: center; color: #111; font-family: 'Inter', sans-serif;">
                Hermes Quant Platform - {report_title}
            </div>
            <div style="font-size: 12px; text-align: right; color: #555; margin-top: 5px; border-bottom: 2px solid #333; padding-bottom: 8px; margin-bottom: 20px; font-family: 'Inter', sans-serif;">
                Report Generated: {current_time_str}
            </div>
        </div>
        <style>
        @media print {{
            .print-only-header {{
                display: block !important;
            }}
            [data-testid="stSidebar"], [data-testid="stHeader"], [data-testid="stToolbar"] {{ display: none !important; }}
            iframe {{ display: none !important; }}
            .stApp {{ background-color: white !important; }}
            * {{ color: black !important; }}
        }}
        </style>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns([5, 1])
    with col2:
        components.html(
            """
            <script>function printDashboard() { window.parent.print(); }</script>
            <div style="display:flex; justify-content:flex-end;">
                <button onclick="printDashboard()" style="background-color: #26A69A; color: white; border: none; padding: 8px 16px; border-radius: 6px; font-family: sans-serif; font-size: 14px; font-weight: bold; cursor: pointer; box-shadow: 0 4px 6px rgba(0,0,0,0.1); transition: all 0.3s ease;">
                    🖨️ Print / Save as PDF
                </button>
            </div>
            <style>
                body { margin: 0; }
                button:hover { transform: scale(1.05) !important; box-shadow: 0 6px 12px rgba(38, 166, 154, 0.4) !important; }
            </style>
            """,
            height=45
        )

st.set_page_config(page_title="Hermes ICT Pro Dashboard", layout="wide", page_icon="📈")

st.markdown("""
<style>
.block-container {
    padding-top: 5rem !important;
}
.fixed-header {
    position: fixed;
    top: 0;
    left: 0;
    right: 0;
    height: 4rem;
    z-index: 99999;
    background-color: #111827 !important; /* Deep dark blue/grey */
    display: flex;
    flex-direction: column;
    justify-content: center;
    align-items: center;
    text-align: center;
    border-bottom: 4px solid #26A69A;
    box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.5);
}
.fixed-header h1 {
    color: #FFFFFF !important;
    margin: 0;
    padding: 0;
    font-size: 1.4rem !important;
    line-height: 1.2;
}
.fixed-header p {
    color: #94A3B8 !important;
    margin: 0;
    padding: 0;
    font-size: 0.75rem !important;
}
[data-testid="stHeader"] {
    background: transparent !important;
    z-index: 100000 !important; /* Keep Streamlit UI elements clickable */
}
/* Ensure the parent container allows sticky elements to work */
.main .block-container {
    overflow: visible;
}
[data-testid="stMetric"] {
    border: 1px solid rgba(255, 255, 255, 0.1);
    border-radius: 8px;
    padding: 10px;
    background: rgba(255, 255, 255, 0.02);
}
[data-testid="stDataFrame"] {
    border: 1px solid rgba(255, 255, 255, 0.1);
    border-radius: 8px;
    padding: 2px;
}
/* Border for Custom Ticker text input in sidebar */
[data-testid="stSidebar"] [data-testid="stTextInput"] div[data-baseweb="input"] {
    border: 2px solid #3B82F6 !important;
    border-radius: 8px !important;
    box-shadow: 0 0 8px rgba(59, 130, 246, 0.4) !important;
}
[data-testid="stSidebar"] [data-testid="stTextInput"] div[data-baseweb="input"]:focus-within {
    border-color: #60A5FA !important;
    box-shadow: 0 0 12px rgba(96, 165, 250, 0.6) !important;
}

/* Sidebar Toggle Icons (keyboard_double_arrow_left/right) */
[data-testid="collapsedControl"] svg,
[data-testid="stSidebarCollapseButton"] svg {
    color: #10B981 !important;
    fill: #10B981 !important;
    stroke: #10B981 !important;
}

/* 📱 Smartphone / Mobile Optimization */
@media (max-width: 768px) {
    .fixed-header {
        position: fixed !important;
        padding-left: 1rem !important;
    }
    .fixed-header h1 {
        font-size: 1.1rem !important;
    }
    .fixed-header p {
        font-size: 0.65rem !important;
    }
    [data-testid="stMetricValue"] {
        font-size: 1.2rem !important;
    }
    /* Scale down all custom HTML spans that use 24px+ */
    span[style*="font-size: 26px"], span[style*="font-size: 24px"], h2[style*="font-size: 32px"], h2[style*="font-size: 48px"] {
        font-size: 20px !important;
    }
    /* Reduce gap in flex containers */
    div[style*="gap: 30px"], div[style*="gap: 20px"] {
        gap: 15px !important;
    }
    /* Adjust padding in custom HTML cards */
    div[style*="padding: 20px"], div[style*="padding: 25px"] {
        padding: 12px !important;
    }
}
</style>
<div class="fixed-header">
    <h1>📈 Hermes ICT Pro Trading Dashboard</h1>
    <p>Institutional Grade Algorithmic Tracking (FVGs, Order Blocks, Liquidity Pools, VWMA).</p>
</div>
""", unsafe_allow_html=True)

# Sidebar Configuration
st.sidebar.header("Configuration")

POPULAR_ASSETS = {
    "AAPL": "Apple Inc.",
    "MSFT": "Microsoft",
    "NVDA": "NVIDIA",
    "TSLA": "Tesla",
    "AMZN": "Amazon",
    "RELIANCE.NS": "Reliance Industries (NSE)",
    "TCS.NS": "Tata Consultancy (NSE)",
    "HDFCBANK.NS": "HDFC Bank (NSE)",
    "INFY.NS": "Infosys (NSE)",
    "SBIN.NS": "State Bank of India (NSE)",
    "EURUSD=X": "EUR/USD Forex",
    "GBPUSD=X": "GBP/USD Forex",
    "JPY=X": "USD/JPY Forex",
    "BTC-USD": "Bitcoin",
    "ETH-USD": "Ethereum",
    "GC=F": "Gold Futures",
    "CL=F": "Crude Oil Futures"
}

SECTOR_INDICES = {
    "^NSEI": "Nifty 50 Index",
    "^NSMIDCP": "Nifty Next 50 Index",
    "^NSEBANK": "Nifty Bank Index",
    "^CNXIT": "Nifty IT Index",
    "^CNXAUTO": "Nifty Auto Index",
    "^CNXFMCG": "Nifty FMCG Index",
    "^CNXMETAL": "Nifty Metal Index",
    "^CNXPHARMA": "Nifty Pharma Index",
    "^CNXREALTY": "Nifty Realty Index",
    "^CNXFIN": "Nifty Financial Services Index",
    "^CNXMEDIA": "Nifty Media Index",
    "^CNXINFRA": "Nifty Infrastructure Index",
    "^NSEMDCP50": "Nifty Midcap 50 Index",
    "^CNXSC": "Nifty Smallcap 100 Index",
    "^CNXENERGY": "Nifty Energy Index",
    "^CNXPSE": "Nifty PSE Index",
    "^CNXMNC": "Nifty MNC Index",
    "^CNXSERVICE": "Nifty Service Sector Index"
}

st.sidebar.markdown("**🔍 Asset Selection**")

@st.cache_data
def load_all_tickers():
    options = []
    # Add popular first
    for k, v in POPULAR_ASSETS.items():
        options.append(f"{k} ({v})")
    # Add sector indices
    for k, v in SECTOR_INDICES.items():
        options.append(f"{k} ({v})")
    # Add NSE
    try:
        import pandas as pd
        df = pd.read_csv("nse_list.csv")
        for _, row in df.iterrows():
            sym = str(row.iloc[0]).strip()
            name = str(row.iloc[1]).strip()
            options.append(f"{sym}.NS ({name})")
    except Exception:
        pass
    return options

all_ticker_options = load_all_tickers()

selected_dropdown = st.sidebar.selectbox("Search Asset (Type to autocomplete):", all_ticker_options, index=1)

st.sidebar.markdown("**Or**")
from streamlit_searchbox import st_searchbox
def search_asset_sidebar(searchterm: str):
    return [o for o in all_ticker_options if searchterm.lower() in o.lower()] if searchterm else []

with st.sidebar:
    custom_ticker = st_searchbox(
        search_asset_sidebar,
        key="sidebar_searchbox",
        placeholder="Enter Custom Ticker..."
    )
    
    st.markdown("---")
    st.markdown("**🎯 Most Probable Setups Settings**")
    scan_interval = st.selectbox(
        "Setups Scan Timeframe", 
        ["1h", "1d"], 
        index=0, 
        key="setups_timeframe_select"
    )
    
    st.markdown("**🤖 Telegram Integration**")
    bot_token = st.text_input(
        "Telegram Bot Token", 
        type="password", 
        placeholder="Enter Bot Token",
        key="setups_telegram_token"
    )
    chat_id = st.text_input(
        "Telegram Chat ID", 
        placeholder="e.g. @sudhir_ict_signals",
        key="setups_telegram_chat"
    )
    
    if st.button("🔌 Test Connection", use_container_width=True):
        if not bot_token or not chat_id:
            st.sidebar.error("Enter Bot Token and Chat ID first!")
        else:
            with st.spinner("Testing..."):
                from telegram_utils import send_telegram_message
                success, resp = send_telegram_message(
                    bot_token, 
                    chat_id, 
                    "🤖 *Hermes AI Engine*\n\nConnection test successful! Telegram integration is fully functional. 🚀"
                )
                if success:
                    st.sidebar.success("✅ Connection Successful!")
                else:
                    st.sidebar.error(resp)

# Initialize session state for ticker persistence to prevent custom searchbox state loss on foreign reruns
if 'active_ticker' not in st.session_state:
    default_ticker = selected_dropdown.split(" ")[0].upper() if selected_dropdown else "AAPL"
    st.session_state['active_ticker'] = default_ticker
    st.session_state['active_ticker_source'] = "dropdown"
    st.session_state['last_dropdown_val'] = default_ticker

# Detect if the custom searchbox returned a value (user typed/selected a custom asset)
if custom_ticker:
    new_ticker = custom_ticker.split(" ")[0].upper()
    if st.session_state.get('active_ticker') != new_ticker:
        st.session_state['active_ticker'] = new_ticker
        st.session_state['active_ticker_source'] = "custom"
else:
    # custom_ticker is None/empty (either cleared or returned None during a rerun of a different tab)
    dropdown_ticker = selected_dropdown.split(" ")[0].upper() if selected_dropdown else "AAPL"
    
    # If the user changed the native selectbox dropdown explicitly, update active ticker
    if st.session_state.get('last_dropdown_val') != dropdown_ticker:
        st.session_state['active_ticker'] = dropdown_ticker
        st.session_state['active_ticker_source'] = "dropdown"
        st.session_state['last_dropdown_val'] = dropdown_ticker

selected_ticker = st.session_state['active_ticker']

@st.cache_data(ttl=60)
def get_stock_data(ticker, period="1y", interval="1d"):
    chart_period = "1mo" if interval == "1h" else period
    df = fetch_data(ticker, period=chart_period, interval=interval)
    df, ict_data = calculate_ict_indicators(ticker, df, generate_logs=True)
    return df, ict_data

def load_logs():
    if os.path.exists(LOG_FILE):
        with open(LOG_FILE, "r") as f:
            try:
                logs = json.load(f)
                return pd.DataFrame(logs)
            except:
                pass
    return pd.DataFrame()

logs_df = load_logs()

# Set up layout tabs
tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs(["📊 Main Dashboard", "📡 Indian Market Scanner", "🎯 Most Probable B/S", "📋 Stock Analysis", "🧭 Sectorial View", "🛢️ Commodities", "📈 Strategy Backtester", "⚡ Custom Scanner"])

# --- TAB 1: MAIN DASHBOARD ---
with tab1:
    chart_interval = "1d"
        
    if ('selected_ticker_data' not in st.session_state or 
        st.session_state.get('last_loaded_ticker') != selected_ticker or 
        st.session_state.get('last_loaded_interval') != chart_interval):
        with st.spinner(f"Loading institutional data for {selected_ticker} ({chart_interval})..."):
            df, ict_data = get_stock_data(selected_ticker, interval=chart_interval)
            st.session_state['selected_ticker_data'] = (df, ict_data)
            st.session_state['last_loaded_ticker'] = selected_ticker
            st.session_state['last_loaded_interval'] = chart_interval
    else:
        df, ict_data = st.session_state['selected_ticker_data']

    add_print_button()

    if not df.empty:
        latest = df.iloc[-1]
        prev = df.iloc[-2]
        
        # ICT HUD
        st.markdown(f"<h2 style='margin-bottom: 0px; margin-top: 0px; color: #3B82F6;'>{selected_ticker}</h2>", unsafe_allow_html=True)
        col1, col2, col3, col4 = st.columns(4)
        current_price = latest['Close']
        price_change = current_price - prev['Close']
        col1.metric("Current Price", f"${current_price:.2f}", f"{price_change:.2f}")
        
        bias_color = "🟢 BULLISH" if ict_data.get('dynamic_bull_bias') else "🔴 BEARISH" if ict_data.get('dynamic_bear_bias') else "⚪ NEUTRAL"
        col2.metric("Structure Shift (Bias)", bias_color)
        col3.metric("RSI Momentum", f"{latest['RSI_14']:.2f}")
        
        latest_signal = "None"
        if not logs_df.empty and 'ticker' in logs_df.columns:
            ticker_logs = logs_df[logs_df['ticker'] == selected_ticker]
            if not ticker_logs.empty:
                latest_signal_row = ticker_logs.iloc[-1]
                latest_signal = f"{latest_signal_row['action']} @ {latest_signal_row['price']}"
        with col4:
            st.caption("Latest ICT Signal")
            st.markdown(f"#### {latest_signal}")
        
        st.markdown("---")
        st.subheader(f"ICT Institutional Analysis - {selected_ticker}")
        
        # Chart Controls aligned in a row
        ctrl_col1, ctrl_col2, ctrl_col3 = st.columns([1, 2, 1])
        with ctrl_col1:
            chart_theme = st.radio("Chart Theme", ["Dark", "Light"], index=1, horizontal=True)
        with ctrl_col2:
            zoom_label = "Chart Zoom (Periods)" if chart_interval == "1h" else "Chart Zoom (Days)"
            max_zoom = min(len(df), 365) if not df.empty else 90
            default_zoom = min(len(df), 90) if not df.empty else 90
            days_to_show = st.slider(zoom_label, min_value=14, max_value=max_zoom, value=default_zoom, step=7)
        with ctrl_col3:
            st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
            refresh = st.button("↻ Refresh Live Data", use_container_width=True)
            if refresh:
                st.cache_data.clear()
        
        fig = go.Figure()
        fig.add_trace(go.Candlestick(x=df.index, open=df['Open'], high=df['High'], low=df['Low'], close=df['Close'], name='OHLC', increasing_line_color='#26A69A', decreasing_line_color='#EF5350', showlegend=False))
        fig.add_trace(go.Scatter(x=df.index, y=df['VWMA_50'], line=dict(color='#FF00FF', width=2), name='VWMA Mean'))
        fig.add_trace(go.Scatter(x=df.index, y=df['CH_High'], line=dict(color='#FFDD00', width=1), name='Channel Cap', opacity=0.7))
        fig.add_trace(go.Scatter(x=df.index, y=df['CH_Low'], line=dict(color='#FFDD00', width=1), name='Channel Floor', opacity=0.7))
        
        max_date = df.index[-1]
        
        for fvg in ict_data.get('fvg', []):
            fcolor = "rgba(0, 255, 170, 0.2)" if fvg['type'] == 'Bull FVG' else "rgba(255, 51, 102, 0.2)"
            end_d = fvg.get('end', max_date)
            fig.add_shape(type="rect", x0=fvg['start'], y0=fvg['bot'], x1=end_d, y1=fvg['top'], line=dict(width=1, color=fcolor.replace("0.2", "0.5")), fillcolor=fcolor, layer="below")
            
        for ob in ict_data.get('ob', []):
            fcolor = "rgba(51, 153, 255, 0.25)" if ob['type'] == 'Bull OB' else "rgba(255, 136, 51, 0.25)"
            end_d = ob.get('end', max_date)
            fig.add_shape(type="rect", x0=ob['start'], y0=ob['bot'], x1=end_d, y1=ob['top'], line=dict(width=1, color=fcolor.replace("0.25", "0.7")), fillcolor=fcolor, layer="below")
            
        for liq in ict_data.get('liq', []):
            lcolor = "#FF5555" if liq['type'] == 'BSL' else "#55FF99"
            fig.add_shape(type="line", x0=liq['index'], y0=liq['price'], x1=max_date, y1=liq['price'], line=dict(color=lcolor, width=2, dash="dash"), layer="below")
            fig.add_annotation(x=max_date, y=liq['price'], text=f" {liq['type']} ", showarrow=False, xanchor="left", font=dict(color=lcolor, size=11))
            
        safe_days = min(days_to_show, len(df)-1)
        zoom_start = df.index[-safe_days] if safe_days > 0 else df.index[0]
        
        theme_template = "plotly_dark" if chart_theme == "Dark" else "plotly_white"
        bg_color = "rgba(0,0,0,0)" # Transparent to show CSS gradient
        grid_color = "rgba(255,255,255,0.05)" if chart_theme == "Dark" else "rgba(0,0,0,0.05)"
        
        # Inject dynamic CSS for chart gradient and border
        st.markdown(f"""
        <style>
        [data-testid="stPlotlyChart"] {{
            background: {'linear-gradient(180deg, #1e293b 0%, #000000 100%)' if chart_theme == 'Dark' else 'linear-gradient(180deg, #ffffff 0%, #e2e8f0 100%)'};
            border: 2px solid {'#334155' if chart_theme == 'Dark' else '#cbd5e1'};
            border-radius: 12px;
            padding: 15px;
            box-shadow: 0 10px 20px rgba(0,0,0,0.2);
        }}
        </style>
        """, unsafe_allow_html=True)

        fig.update_layout(
            xaxis_rangeslider_visible=False,
            template=theme_template, 
            height=800, 
            plot_bgcolor=bg_color, 
            paper_bgcolor=bg_color, 
            margin=dict(l=0, r=60, t=20, b=0), 
            xaxis=dict(
                showgrid=True, gridcolor=grid_color, griddash='dot', 
                rangebreaks=[dict(bounds=["sat", "mon"])],
                range=[zoom_start, max_date],
                showspikes=True, spikemode='across', spikethickness=1, spikedash='dot', spikecolor='#999999'
            ), 
            yaxis=dict(
                side='right',
                showgrid=True, gridcolor=grid_color, griddash='dot', 
                tickformat='.2f',
                showspikes=True, spikemode='across', spikethickness=1, spikedash='dot', spikecolor='#999999'
            ), 
            legend=dict(
                orientation="v", yanchor="top", y=0.99, xanchor="left", x=0.01, 
                bgcolor="rgba(0,0,0,0)"
            ),
            hovermode='x unified'
        )
        st.plotly_chart(fig, use_container_width=True)
        
        st.markdown("---")
        st.subheader("Institutional Triggers Log")
        if not logs_df.empty:
            display_df = logs_df.iloc[::-1].reset_index(drop=True)
            st.dataframe(display_df, use_container_width=True)
        else:
            st.info("No active decisions logged yet.")

        st.markdown("---")
        st.subheader("📖 Understanding the Institutional Triggers")
        st.markdown("The triggers log captures key algorithmic actions. Here is what they mean for your trading strategy:")

        col_exp1, col_exp2, col_exp3, col_exp4 = st.columns(4)
        with col_exp1:
            st.info("**MSS (Market Structure Shift)**\n\nIndicates a reversal in trend bias. When price violently breaks a major recent swing high or low, it signals institutions are reversing the market direction.")
        with col_exp2:
            st.success("**FVG (Fair Value Gap)**\n\nA rapid 3-bar pricing inefficiency. Institutions moved so fast they left a gap. Price often gravitates back to this zone to 'fill' or 'mitigate' it before continuing.")
        with col_exp3:
            st.warning("**Order Blocks (OB)**\n\nThe exact candlestick footprint where banks accumulated massive positions (detected via high ATR displacement). It acts as a powerful future support/resistance magnet.")
        with col_exp4:
            st.error("**Liquidity Pools (BSL/SSL)**\n\nAreas where retail stop-losses rest (above old highs and below old lows). Algorithms frequently push price into these lines to grab liquidity before reversing.")

        st.markdown("### The Standard Institutional Trade Flow")
        funnel_fig = go.Figure(go.Funnelarea(
            text=["1. Liquidity Sweep (BSL/SSL)", "2. Structure Shift (MSS)", "3. Footprint Left (OB/FVG)", "4. Retracement & Mitigation", "5. Institutional Expansion (Take Profit)"],
            values=[100, 80, 60, 40, 20],
            marker={"colors": ["#FF5555", "#FFDD00", "#3399FF", "#00FFAA", "#BB77FF"]},
            textfont={"color": "black", "size": 14}
        ))
        funnel_fig.update_layout(template=theme_template if 'theme_template' in locals() else "plotly_dark", paper_bgcolor=bg_color if 'bg_color' in locals() else "#0e1117", height=350, margin=dict(t=30, b=0, l=0, r=0))
        st.plotly_chart(funnel_fig, use_container_width=True)

        st.markdown("---")
        st.subheader(f"🔮 Automated Forecast & Key Levels ({selected_ticker})")
        
        # S/R Logic
        key_levels = []
        for liq in ict_data.get('liq', []): key_levels.append({'name': liq['type'] + " (Liquidity)", 'price': liq['price']})
        for ob in ict_data.get('ob', []):
            if ob['type'] == 'Bull OB': key_levels.append({'name': 'Bull OB Support', 'price': ob['top']})
            else: key_levels.append({'name': 'Bear OB Resistance', 'price': ob['bot']})
        for fvg in ict_data.get('fvg', []):
            if fvg.get('active', True):
                if fvg['type'] == 'Bull FVG': key_levels.append({'name': 'Bull FVG Support', 'price': fvg['top']})
                else: key_levels.append({'name': 'Bear FVG Resistance', 'price': fvg['bot']})

        unique_levels = []
        seen_prices = set()
        for lvl in key_levels:
            p_rounded = round(lvl['price'], 2)
            if p_rounded not in seen_prices:
                seen_prices.add(p_rounded)
                unique_levels.append(lvl)

        supports = sorted([lvl for lvl in unique_levels if lvl['price'] < current_price], key=lambda x: x['price'], reverse=True)
        resistances = sorted([lvl for lvl in unique_levels if lvl['price'] > current_price], key=lambda x: x['price'])

        is_bull = ict_data.get('dynamic_bull_bias', False)
        is_bear = ict_data.get('dynamic_bear_bias', False)
        forecast_text = "Neutral - Waiting for clear Market Structure Shift."
        forecast_color = "gray"

        if is_bull:
            if latest['RSI_14'] > 70:
                forecast_text = "⚠️ **Bullish Overextended:** Momentum is high, but RSI is overbought. A temporary retracement down to the nearest Support is highly probable before further upside."
                forecast_color = "#FFDD00"
            else:
                forecast_text = "🚀 **Bullish Continuation:** Market structure is intact with healthy momentum. Price is projected to hunt the nearest Resistance levels (Buy-Side Liquidity)."
                forecast_color = "#00FFAA"
        elif is_bear:
            if latest['RSI_14'] < 30:
                forecast_text = "⚠️ **Bearish Exhausted:** Momentum is heavy, but RSI is oversold. Expect a short-term relief rally up to the nearest Resistance before continuing the downtrend."
                forecast_color = "#FFDD00"
            else:
                forecast_text = "📉 **Bearish Continuation:** Market structure is heavily bearish. Price is projected to hunt lower Support levels (Sell-Side Liquidity)."
                forecast_color = "#FF3366"

        text_color = "#000000" # Forced to black as requested
        st.markdown(f"<div style='padding:15px; border-radius:5px; border-left: 5px solid {forecast_color}; background-color: rgba(255,255,255,0.9); color: {text_color}; font-size: 16px; margin-bottom: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);'>{forecast_text}</div>", unsafe_allow_html=True)

        col_sup, col_res = st.columns(2)
        with col_sup:
            st.markdown("🛡️ **Immediate Support Levels (Demand)**")
            if supports:
                for i, s in enumerate(supports[:3]): st.markdown(f"- **${s['price']:.2f}** ➔ {s['name']}")
            else: st.markdown("- *No immediate structural support found nearby.*")

        with col_res:
            st.markdown("🎯 **Immediate Resistance Levels (Supply)**")
            if resistances:
                for i, r in enumerate(resistances[:3]): st.markdown(f"- **${r['price']:.2f}** ➔ {r['name']}")
            else: st.markdown("- *No immediate structural resistance found nearby.*")

    else:
        st.warning("No data available for the selected ticker.")

with tab2:
    st.subheader("📡 Interactive Paginated Market Scanner (2000+ Stocks)")
    st.markdown("Select a page below to instantly scan a block of 100 stocks. The algorithmic engine will use 20 parallel threads to return your institutional bias results in roughly 10-15 seconds.")
    
    @st.cache_data
    def get_all_nse_tickers():
        if not os.path.exists("nse_list.csv"): return []
        try:
            df_nse = pd.read_csv("nse_list.csv")
            if 'SYMBOL' in df_nse.columns:
                return [str(sym).strip() + ".NS" for sym in df_nse['SYMBOL'].tolist() if pd.notna(sym)]
        except Exception: return []
        return []

    all_tickers = get_all_nse_tickers()
    total_tickers = len(all_tickers)
    
    if total_tickers == 0:
        st.warning("Could not load nse_list.csv. Please ensure the master file exists.")
    else:
        PAGE_SIZE = 100
        total_pages = (total_tickers // PAGE_SIZE) + (1 if total_tickers % PAGE_SIZE > 0 else 0)
        
        col_page, col_btn = st.columns([1, 2])
        with col_page:
            selected_page = st.number_input(f"Select Page (1 to {total_pages})", min_value=1, max_value=total_pages, value=1)
            
        start_idx = (selected_page - 1) * PAGE_SIZE
        end_idx = min(start_idx + PAGE_SIZE, total_tickers)
        current_batch = all_tickers[start_idx:end_idx]
        
        st.markdown(f"**Ready to scan stocks {start_idx + 1} to {end_idx} (out of {total_tickers})**")
        
        if st.button(f"🚀 Scan Page {selected_page} Now", use_container_width=True):
            bullish_list = []
            bearish_list = []
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            from concurrent.futures import ThreadPoolExecutor, as_completed
            
            def scan_single(ticker):
                try:
                    df_s = fetch_data(ticker, period="6mo", interval="1d")
                    if df_s.empty or len(df_s) < 50: return None
                    df_s, ict_s = calculate_ict_indicators(ticker, df_s, generate_logs=False)
                    cur_price = df_s.iloc[-1]['Close']
                    cur_rsi = df_s.iloc[-1]['RSI_14']
                    bull = ict_s.get('dynamic_bull_bias', False)
                    bear = ict_s.get('dynamic_bear_bias', False)
                    if bull or bear:
                        return {"Ticker": ticker.replace(".NS", ""), "Price (₹)": round(cur_price, 2), "RSI": round(cur_rsi, 2), "Bias": "Bullish" if bull else "Bearish"}
                except Exception: return None
                return None

            completed = 0
            with ThreadPoolExecutor(max_workers=20) as executor:
                future_to_ticker = {executor.submit(scan_single, t): t for t in current_batch}
                for future in as_completed(future_to_ticker):
                    completed += 1
                    status_text.text(f"Scanning... ({completed}/{len(current_batch)})")
                    progress_bar.progress(completed / len(current_batch))
                    
                    res = future.result()
                    if res:
                        if res['Bias'] == 'Bullish': bullish_list.append(res)
                        else: bearish_list.append(res)
                            
            status_text.text(f"Page {selected_page} Scan Complete!")
            
            col_bull, col_bear = st.columns(2)
            with col_bull:
                st.success(f"🟢 **Bullish Bias Detected ({len(bullish_list)})**")
                if bullish_list:
                    bullish_list = sorted(bullish_list, key=lambda x: x['RSI'], reverse=True)
                    st.dataframe(pd.DataFrame(bullish_list).drop(columns=['Bias']), use_container_width=True, hide_index=True)
                else: st.info("No bullish structures found on this page.")
                    
            with col_bear:
                st.error(f"🔴 **Bearish Bias Detected ({len(bearish_list)})**")
                if bearish_list:
                    bearish_list = sorted(bearish_list, key=lambda x: x['RSI'])
                    st.dataframe(pd.DataFrame(bearish_list).drop(columns=['Bias']), use_container_width=True, hide_index=True)
                else: st.info("No bearish structures found on this page.")

with tab3:
    st.subheader("🎯 Top 10 Most Probable Bullish & Bearish Setups")
    st.markdown("This engine cross-references the **ICT Technical Scanner** with the **Hermes AI Fundamental Engine** to find the absolute highest probability trades.")
    
    with st.expander("📖 View Overlapping Confluence Conditions (Rules)", expanded=False):
        col_r1, col_r2 = st.columns(2)
        with col_r1:
            st.markdown("""
            ### 🟢 Bullish Entry Confluence (Buy Setup)
            To execute a perfect buy order, the system requires the following overlapping conditions:
            * **HTF Confluence (4H EMA):** Price must close above the 4H EMA (the bullish gate) to align with the major trend.
            * **Market Structure Shift (MSS):** The background/bias must be Green, indicating a Bullish MSS (candle body closed beyond the last swing high).
            * **Liquidity Sweep:** Price must sweep below the Green Dashed SSL line at a swing low before reversing (low of the last 5 bars was <= SSL).
            * **Institutional Order Block (OB):** Price respects or reacts to a Blue Box Bull OB (latest close is above OB bottom and <= top * 1.05).
            * **Fair Value Gap (FVG):** An active Green Box Bull FVG must form on the chart.
            """)
        with col_r2:
            st.markdown("""
            ### 🔴 Bearish Entry Confluence (Sell Setup)
            To execute a perfect sell order, the system requires the following overlapping conditions:
            * **HTF Confluence (4H EMA):** Price must close below the 4H EMA (the bearish gate) to filter out counter-trend noise.
            * **Market Structure Shift (MSS):** The background/bias must be Red, indicating a Bearish MSS (candle body closed beyond the last swing low).
            * **Liquidity Sweep:** Price must sweep above the Red Dashed BSL line at a swing high before reversing (high of the last 5 bars was >= BSL).
            * **Institutional Order Block (OB):** Price respects or reacts to an Orange Box Bear OB (latest close is below OB top and >= bot * 0.95).
            * **Fair Value Gap (FVG):** An active Red Box Bear FVG must form on the chart.
            """)
            
    st.info("💡 **Setup Scanner Settings**: Configure your timeframe and Telegram Bot credentials in the sidebar menu on the left.")
    
    st.markdown("---")
    
    col_btn1, col_btn2, col_btn3 = st.columns([2, 2, 1])
    with col_btn1:
        analyze_clicked = st.button("🚀 Analyze Top 10 B/S Setups", use_container_width=True)
    with col_btn2:
        broadcast_clicked = st.button("📤 Broadcast to Telegram", use_container_width=True)
    
    if analyze_clicked:
        if not os.path.exists("scanner_results.json"):
            st.warning("No scanner results found! Please run the 'Indian Market Scanner' on at least one page first.")
        else:
            with st.spinner("Crunching fundamental data for the Top 20 ICT Setups... This takes about 10 seconds..."):
                try:
                    with open("scanner_results.json", "r") as f:
                        scan_data = json.load(f)
                    
                    bullish_raw = scan_data.get("bullish", [])
                    bearish_raw = scan_data.get("bearish", [])
                    
                    from fundamental import get_fundamental_score
                    
                    from concurrent.futures import ThreadPoolExecutor, as_completed
                    
                    def process_single_candidate(item, bias):
                        try:
                            # Filter 1: Price > 55
                            if item.get("Price", 0) <= 55:
                                return None
                                
                            ticker = item["Ticker"]
                            if not ticker.endswith(".NS") and not ticker.endswith(".BO"):
                                ticker += ".NS"
                                
                            score, verdict, volume = get_fundamental_score(ticker)
                            
                            # Filter 2: Volume > 300,000
                            if volume <= 300000:
                                return None
                                
                            # Fetch data based on timeframe
                            scan_period = "3mo" if scan_interval == "1h" else "1y"
                            df_tf = fetch_data(ticker, period=scan_period, interval=scan_interval)
                            if df_tf.empty or len(df_tf) < 50:
                                return None
                                
                            df_tf, ict_tf = calculate_ict_indicators(ticker, df_tf, generate_logs=False)
                            if df_tf.empty:
                                return None
                                
                            latest_row = df_tf.iloc[-1]
                            latest_close = latest_row['Close']
                            obs = ict_tf.get('ob', [])
                            
                            # Calculate 4H EMA
                            if scan_interval == "1h":
                                df_1h = df_tf
                            else:
                                df_1h = fetch_data(ticker, period="3mo", interval="1h")
                                
                            if df_1h.empty or len(df_1h) < 100:
                                return None
                                
                            # Resample to 4H
                            ohlc_dict = {
                                'Open': 'first',
                                'High': 'max',
                                'Low': 'min',
                                'Close': 'last',
                                'Volume': 'sum'
                            }
                            agg_dict = {col: ohlc_dict[col] for col in df_1h.columns if col in ohlc_dict}
                            try:
                                df_4h = df_1h.resample('4h', origin='start').agg(agg_dict).dropna()
                            except TypeError:
                                df_4h = df_1h.resample('4h').agg(agg_dict).dropna()
                                
                            if len(df_4h) < 50:
                                return None
                            df_4h['EMA_50'] = df_4h['Close'].ewm(span=50, adjust=False).mean()
                            latest_4h_ema = df_4h['EMA_50'].iloc[-1]
                            
                            if bias == "Bullish":
                                # 1. Higher Timeframe Bias (HTF Confluence): Price must close above the 4H EMA
                                if latest_close <= latest_4h_ema:
                                    return None
                                    
                                # 2. Market Structure Shift (MSS): Background must turn Green, indicating a Bullish MSS
                                if not ict_tf.get('dynamic_bull_bias', False):
                                    return None
                                    
                                # 3. Liquidity Sweep: Price should ideally sweep below the Green Dashed + SSL line
                                ssl_lines = [liq for liq in ict_tf.get('liq', []) if liq['type'] == 'SSL']
                                if not ssl_lines:
                                    return None
                                latest_ssl = ssl_lines[-1]['price']
                                recent_lows = df_tf['Low'].iloc[-5:].values
                                if not any(l <= latest_ssl for l in recent_lows):
                                    return None
                                    
                                # 4. Institutional Order Block (OB): Price should respect or react to a Blue Box
                                bull_obs = [ob for ob in obs if ob['type'] == 'Bull OB']
                                if not bull_obs:
                                    return None
                                most_recent_ob = bull_obs[-1]
                                if not most_recent_ob.get('active', True):
                                    return None
                                if latest_close <= most_recent_ob['bot'] or latest_close > most_recent_ob['top'] * 1.05:
                                    return None
                                    
                                # 5. Fair Value Gap (FVG): A Green Box must form
                                bull_fvgs = [fvg for fvg in ict_tf.get('fvg', []) if fvg['type'] == 'Bull FVG' and fvg.get('active', True)]
                                if not bull_fvgs:
                                    return None
                                    
                            elif bias == "Bearish":
                                # 1. Higher Timeframe Bias (HTF Confluence): Price must close below the 4H EMA
                                if latest_close >= latest_4h_ema:
                                    return None
                                    
                                # 2. Market Structure Shift (MSS): Background must turn Red, indicating a Bearish MSS
                                if not ict_tf.get('dynamic_bear_bias', False):
                                    return None
                                    
                                # 3. Liquidity Sweep: Price should ideally sweep above the Red Dashed + BSL line
                                bsl_lines = [liq for liq in ict_tf.get('liq', []) if liq['type'] == 'BSL']
                                if not bsl_lines:
                                    return None
                                latest_bsl = bsl_lines[-1]['price']
                                recent_highs = df_tf['High'].iloc[-5:].values
                                if not any(h >= latest_bsl for h in recent_highs):
                                    return None
                                    
                                # 4. Institutional Order Block (OB): Price should respect or react to an Orange Box
                                bear_obs = [ob for ob in obs if ob['type'] == 'Bear OB']
                                if not bear_obs:
                                    return None
                                most_recent_ob = bear_obs[-1]
                                if not most_recent_ob.get('active', True):
                                    return None
                                if latest_close >= most_recent_ob['top'] or latest_close < most_recent_ob['bot'] * 0.95:
                                    return None
                                    
                                # 5. Fair Value Gap (FVG): A Red Box must form
                                bear_fvgs = [fvg for fvg in ict_tf.get('fvg', []) if fvg['type'] == 'Bear FVG' and fvg.get('active', True)]
                                if not bear_fvgs:
                                    return None
                                    
                            verdict_clean = verdict.replace("Strong ", "")
                            v_emoji = "🟢" if "BUY" in verdict_clean or "Accumulate" in verdict_clean else "🔴" if "SELL" in verdict_clean else "⚪"
                            
                            return {
                                "Ticker": item["Ticker"],
                                "Price": item.get("Price", "N/A"),
                                "ICT Bias": item["Bias"],
                                "RSI": item["RSI"],
                                "Vol (K)": f"{volume/1000:.0f}K",
                                "Fund. Score": f"{score}/10",
                                "AI Verdict": f"{v_emoji} {verdict}"
                            }
                        except Exception:
                            return None

                    def enrich_list(raw_list, bias, req_count=10):
                        # Filter out cheap stocks first
                        filtered_raw = [item for item in raw_list if item.get("Price", 0) > 55][:40]
                        
                        enriched = []
                        with ThreadPoolExecutor(max_workers=8) as executor:
                            futures = {executor.submit(process_single_candidate, item, bias): item for item in filtered_raw}
                            for future in as_completed(futures):
                                res = future.result()
                                if res is not None:
                                    enriched.append(res)
                                    
                        # Sort to maintain original RSI order
                        ticker_to_index = {item["Ticker"]: i for i, item in enumerate(filtered_raw)}
                        enriched = sorted(enriched, key=lambda x: ticker_to_index.get(x["Ticker"], 999))
                        
                        return enriched[:req_count]
                        
                    bullish_final = enrich_list(bullish_raw, "Bullish", 10)
                    bearish_final = enrich_list(bearish_raw, "Bearish", 10)
                        
                    st.session_state['bullish_final'] = bullish_final
                    st.session_state['bearish_final'] = bearish_final
                    
                except Exception as e:
                    st.error(f"An error occurred during analysis: {e}")

    # Display results if they exist in session state
    if 'bullish_final' in st.session_state and 'bearish_final' in st.session_state:
        bullish_final = st.session_state['bullish_final']
        bearish_final = st.session_state['bearish_final']
        
        col_b1, col_b2 = st.columns(2)
        with col_b1:
            st.success(f"📈 **Top {len(bullish_final)} Bullish Candidates**")
            if bullish_final:
                st.dataframe(pd.DataFrame(bullish_final), use_container_width=True, hide_index=True)
            else:
                st.info("No bullish candidates found in scanner results.")
                
        with col_b2:
            st.error(f"📉 **Top {len(bearish_final)} Bearish Candidates**")
            if bearish_final:
                st.dataframe(pd.DataFrame(bearish_final), use_container_width=True, hide_index=True)
            else:
                st.info("No bearish candidates found in scanner results.")
                
        # --- PDF Download Generation ---
        def generate_pdf_bytes(bullish, bearish):
            import tempfile, os
            try:
                from fpdf import FPDF
            except ImportError:
                return None
            from datetime import datetime
            current_date = datetime.now().strftime("%B %d, %Y | %H:%M")

            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(190, 10, txt="Hermes ICT Pro - Top 10 B/S Setups", ln=True, align='C')
            
            pdf.set_font("Arial", 'I', 10)
            pdf.set_text_color(100, 100, 100) # grey color
            pdf.cell(190, 6, txt=f"Generated on: {current_date}", ln=True, align='C')
            pdf.set_text_color(0, 0, 0) # reset
            pdf.ln(5)
            def draw_table(title, data, is_bullish):
                pdf.set_font("Arial", 'B', 14)
                if is_bullish:
                    pdf.set_text_color(0, 128, 0) # Dark Green
                else:
                    pdf.set_text_color(200, 0, 0) # Dark Red
                    
                pdf.cell(190, 10, txt=title, ln=True, align='L')
                pdf.set_text_color(0, 0, 0) # Reset to black
                pdf.ln(2)
                
                # Table Headers
                pdf.set_font("Arial", 'B', 10)
                pdf.set_fill_color(240, 240, 240)
                headers = ['Ticker', 'Price', 'RSI', 'Vol (K)', 'Score', 'Verdict']
                widths = [35, 25, 20, 25, 20, 65]
                
                for i in range(len(headers)):
                    pdf.cell(widths[i], 8, txt=headers[i], border=1, align='C', fill=True)
                pdf.ln()
                
                # Table Rows
                pdf.set_font("Arial", '', 10)
                for row in data:
                    pdf.cell(widths[0], 8, txt=str(row.get('Ticker', '')), border=1, align='C')
                    pdf.cell(widths[1], 8, txt=str(row.get('Price', '')), border=1, align='C')
                    pdf.cell(widths[2], 8, txt=str(row.get('RSI', '')), border=1, align='C')
                    pdf.cell(widths[3], 8, txt=str(row.get('Vol (K)', '')), border=1, align='C')
                    pdf.cell(widths[4], 8, txt=str(row.get('Fund. Score', '')), border=1, align='C')
                    
                    verdict = str(row.get('AI Verdict', ''))
                    verdict = verdict.encode('ascii', 'ignore').decode('ascii').strip()
                    pdf.cell(widths[5], 8, txt=verdict, border=1, align='C')
                    pdf.ln()
                pdf.ln(10)

            draw_table("Top Bullish Candidates", bullish, True)
            draw_table("Top Bearish Candidates", bearish, False)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                pdf.output(tmp.name)
                with open(tmp.name, "rb") as f: data = f.read()
            os.unlink(tmp.name)
            return data

        pdf_data = generate_pdf_bytes(bullish_final, bearish_final)
        with col_btn3:
            if pdf_data:
                st.download_button(label="📥 Download PDF", data=pdf_data, file_name="Hermes_Top_10_Setups.pdf", mime="application/pdf", use_container_width=True)
            else:
                st.button("📥 Download PDF", disabled=True, help="Installing PDF tools in background...", use_container_width=True)

        if broadcast_clicked:
            if not bot_token or not chat_id:
                st.error("Please enter your Telegram Bot Token and Chat ID to broadcast.")
            else:
                from telegram_utils import format_telegram_message, send_telegram_message
                msg_text = format_telegram_message(bullish_final, bearish_final)
                success, resp = send_telegram_message(bot_token, chat_id, msg_text)
                if success:
                    st.balloons()
                    st.success("✅ " + resp)
                else:
                    st.error("❌ " + resp)

with tab4:
    st.subheader("📋 Expert Fundamental Analyst")
    st.markdown("Act as an expert equity research analyst. Enter any stock ticker below to conduct a rigorous, data-driven fundamental analysis using live valuation, profitability, and health metrics.")
    
    col_input, col_btn2 = st.columns([2, 1])
    with col_input:
        analysis_ticker_raw = st.selectbox("Search Asset (Type to autocomplete):", all_ticker_options, index=0)
        st.markdown("**Or**")
        from streamlit_searchbox import st_searchbox
        def search_asset_expert(searchterm: str):
            return [o for o in all_ticker_options if searchterm.lower() in o.lower()] if searchterm else []
            
        custom_analysis_ticker = st_searchbox(
            search_asset_expert,
            key="expert_searchbox",
            placeholder="Enter Custom Ticker..."
        )
        
        if custom_analysis_ticker:
            analysis_ticker = custom_analysis_ticker.split(" ")[0].upper()
        else:
            analysis_ticker = analysis_ticker_raw.split(" ")[0] if analysis_ticker_raw else "RELIANCE.NS"
    
    st.markdown("""
        <style>
        /* Add a sleek hover animation to buttons */
        div.stButton > button {
            transition: all 0.3s ease-in-out;
            border-radius: 8px;
        }
        div.stButton > button:hover {
            transform: scale(1.05);
            box-shadow: 0px 5px 15px rgba(38, 166, 154, 0.4);
            border-color: #26A69A;
        }
        </style>
    """, unsafe_allow_html=True)
    
    if st.button("Generate Expert Report"):
        if analysis_ticker:
            with st.spinner(f"Analyzing {analysis_ticker}... fetching live fundamental data from Yahoo Finance..."):
                from fundamental import generate_fundamental_report
                report_md = generate_fundamental_report(analysis_ticker.strip().upper())
                add_print_button()
                st.markdown(report_md, unsafe_allow_html=True)
        else:
            st.warning("Please enter a ticker symbol.")

with tab5:
    st.subheader("🧭 Hermes AI Sectorial View")
    st.markdown("Visualizing the flow of institutional liquidity across major Indian sectors. Colored from **Pastel Green (Trending)** to **Pastel Red (Distribution)** based on the Hermes Trend Score.")
    
    if st.button("Generate Sector Heatmap", use_container_width=True):
        with st.spinner("Analyzing top 10 NSE Sectors... This pulls live data and may take 5-10 seconds."):
            try:
                from engine import analyze_indian_sectors
                sector_data = analyze_indian_sectors()
                if sector_data:
                    st.session_state['sector_data'] = sector_data
                else:
                    st.error("Failed to fetch sector data from Yahoo Finance.")
            except Exception as e:
                st.error(f"Error fetching data: {e}")

    if 'sector_data' in st.session_state:
        try:
            add_print_button(report_title="Professional Sectorial Report")
            import plotly.express as px
            import pandas as pd
            
            df_sectors = pd.DataFrame(st.session_state['sector_data'])
            df_sectors['SliceSize'] = 1
            
            # Use bar_polar since pie does not support color_continuous_scale
            fig = px.bar_polar(
                df_sectors,
                r='SliceSize',
                theta='Sector',
                color='TrendScore',
                color_continuous_scale='RdYlGn',
                hover_data=["TrendScore", "Return_1M", "RSI"]
            )
            
            fig.update_traces(
                hovertemplate="<b>%{theta}</b><br>Hermes Score: %{customdata[0]:.2f}<br>1M Return: %{customdata[1]:.2f}%<br>RSI: %{customdata[2]:.2f}<extra></extra>"
            )
            
            fig.update_layout(
                polar=dict(
                    radialaxis=dict(showticklabels=False, ticks='', showgrid=False),
                    angularaxis=dict(tickfont=dict(size=14, color='#000000')),
                    hole=0.4
                ),
                showlegend=False,
                margin=dict(t=20, b=20, l=20, r=20),
                paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(0,0,0,0)",
                coloraxis_colorbar=dict(title="Trend Score")
            )
            
            st.plotly_chart(fig, use_container_width=True)
            
            st.markdown("### 📊 Raw Sector Metrics")
            
            # Color coding the dataframe
            def color_score(val):
                if val >= 15: return 'color: #2ecc71; font-weight: bold;' # Light Green
                elif val >= 0: return 'color: #27ae60;' # Dark Green
                elif val <= -15: return 'color: #e74c3c; font-weight: bold;' # Light Red
                else: return 'color: #c0392b;' # Dark Red
                
            def get_star_rating(score):
                if score >= 15: return '⭐⭐⭐⭐⭐'
                elif score >= 5: return '⭐⭐⭐⭐'
                elif score >= -5: return '⭐⭐⭐'
                elif score >= -15: return '⭐⭐'
                else: return '⭐'
                
            display_df = df_sectors[['Sector', 'TrendScore', 'Return_1M', 'RSI']].sort_values('TrendScore', ascending=False)
            display_df['Star Ratings'] = display_df['TrendScore'].apply(get_star_rating)
            
            # Reorder columns
            display_df = display_df[['Sector', 'Star Ratings', 'TrendScore', 'Return_1M', 'RSI']]
            
            # Support both Pandas >= 2.1.0 (has .map) and older versions (has .applymap)
            styler = display_df.style
            if hasattr(styler, 'map'):
                styler = styler.map(color_score, subset=['TrendScore'])
            else:
                styler = styler.applymap(color_score, subset=['TrendScore'])

            styled_df = (styler
                .format({'TrendScore': "{:.2f}", 'Return_1M': "{:.2f}%", 'RSI': "{:.2f}"})
                .set_properties(**{'text-align': 'right'})
            )
            st.dataframe(styled_df, hide_index=True, use_container_width=False)

            # Intelligent Summary
            st.markdown("### 🧠 Hermes AI Sector Intelligence")
            for _, row in display_df.iterrows():
                sector = row['Sector']
                score = row['TrendScore']
                ret = row['Return_1M']
                rsi = row['RSI']
                
                
                # Custom CSS styling for the cards
                base_style = "padding: 15px; border-radius: 8px; box-shadow: 0px 4px 12px rgba(0,0,0,0.3); margin-bottom: 15px; border: 1px solid rgba(255,255,255,0.05);"
                
                if score >= 15:
                    verdict = "🟢 STRONGLY BULLISH"
                    summary = f"Institutional liquidity is aggressively flowing into **{sector}**. With an impressive **{ret:.2f}%** monthly return and high momentum (RSI: **{rsi:.1f}**), this sector is clearly leading the broader market. Pullbacks present high-probability buying opportunities."
                    card_style = f"{base_style} border-left: 5px solid #2ecc71; background-color: rgba(46, 204, 113, 0.1);"
                elif score >= 0:
                    verdict = "🟢 ACCUMULATION"
                    summary = f"**{sector}** is showing positive underlying strength. A modest **{ret:.2f}%** return combined with stable momentum suggests steady institutional accumulation. The sector is absorbing supply and preparing for continuation."
                    card_style = f"{base_style} border-left: 5px solid #3498db; background-color: rgba(52, 152, 219, 0.1);"
                elif score <= -15:
                    verdict = "🔴 STRONGLY BEARISH"
                    summary = f"Severe distribution detected in **{sector}**. A harsh **{ret:.2f}%** drop and deeply weak momentum (RSI: **{rsi:.1f}**) indicates heavy institutional offloading. Avoid long positions until structural breaks are observed."
                    card_style = f"{base_style} border-left: 5px solid #e74c3c; background-color: rgba(231, 76, 60, 0.1);"
                else:
                    verdict = "🔴 DISTRIBUTION / CHOP"
                    summary = f"**{sector}** is currently facing headwinds. Negative returns (**{ret:.2f}%**) and lagging momentum suggest the sector is out of favor. Liquidity is likely rotating out of this sector and into stronger ones."
                    card_style = f"{base_style} border-left: 5px solid #f1c40f; background-color: rgba(241, 196, 15, 0.1);"

                html_content = f'''
                <div style="{card_style}">
                    <h4 style="margin-top: 0px; margin-bottom: 10px; font-size: 16px;">{verdict} | {sector} <span style="font-size:14px; opacity: 0.7;">(Score: {score:.1f})</span></h4>
                    <p style="margin-bottom: 0px; font-size: 14px; line-height: 1.5;">{summary}</p>
                </div>
                '''
                st.markdown(html_content, unsafe_allow_html=True)

        except Exception as e:
            st.error(f"Error generating sectorial view: {e}")

with tab6:
    st.subheader("🛢️ Commodities Expert Signals (ICT)")
    st.markdown("Live algorithmic analysis for Gold, Silver, Crude Oil, and Natural Gas using Smart Money Concepts (FVG, Order Blocks) to map institutional liquidity zones.")
    
    commodities = {
        "Gold": "GC=F",
        "Silver": "SI=F",
        "Crude Oil (WTI)": "CL=F",
        "Natural Gas": "NG=F"
    }
    
    if st.button("Generate Expert Commodities Report", use_container_width=True):
        add_print_button(report_title="Professional Commodities Report")
        with st.spinner("Scanning Global Commodities Data..."):
            for comm_name, comm_ticker in commodities.items():
                try:
                    df_comm, ict_data = get_stock_data(comm_ticker, period="6mo")
                except Exception as e:
                    st.error(f"Failed to fetch data for {comm_name} ({comm_ticker}).")
                    continue
                    
                if df_comm is None or df_comm.empty:
                    st.error(f"Failed to fetch data for {comm_name} ({comm_ticker}).")
                    continue
                
                current_price = df_comm['Close'].iloc[-1]
                
                fvgs = ict_data.get('fvg', [])
                obs = ict_data.get('ob', [])
                
                buy_price = 0
                for fvg in fvgs:
                    if fvg['type'] == 'Bull FVG' and fvg['top'] < current_price:
                        if fvg['top'] > buy_price: buy_price = fvg['top']
                
                for ob in obs:
                    if ob['type'] == 'Bull OB' and ob['top'] < current_price:
                        if ob['top'] > buy_price: buy_price = ob['top']
                        
                sell_price = float('inf')
                for fvg in fvgs:
                    if fvg['type'] == 'Bear FVG' and fvg['bot'] > current_price:
                        if fvg['bot'] < sell_price: sell_price = fvg['bot']
                
                for ob in obs:
                    if ob['type'] == 'Bear OB' and ob['bot'] > current_price:
                        if ob['bot'] < sell_price: sell_price = ob['bot']
                        
                if sell_price == float('inf'):
                    sell_price = current_price * 1.05 # default resistance
                if buy_price == 0:
                    buy_price = current_price * 0.95
                    
                trend_verdict = "🟢 STRONGLY BULLISH" if current_price > df_comm['Close'].rolling(50).mean().iloc[-1] else "🔴 STRONGLY BEARISH"
                
                buy_dist_pct = ((current_price - buy_price) / current_price) * 100 if current_price else 0
                sell_dist_pct = ((sell_price - current_price) / current_price) * 100 if current_price else 0
                rr_ratio = (sell_price - current_price) / (current_price - buy_price) if (current_price - buy_price) > 0 else 0

                if trend_verdict.startswith("🟢"):
                    bias_desc = "Primary institutional bias remains <b style='color:#2ecc71;'>BULLISH</b>. Buy-side liquidity is in control."
                    action_desc = "Wait for a retracement into the discount array."
                else:
                    bias_desc = "Primary institutional bias remains <b style='color:#e74c3c;'>BEARISH</b>. Sell programs are dominating the tape."
                    action_desc = "Seek premium arrays for high-probability short entries."
                    
                # Render UI box for the commodity
                base_style = "padding: 20px; border-radius: 10px; box-shadow: 0px 8px 24px rgba(0,0,0,0.4); margin-bottom: 20px; border: 1px solid rgba(255,255,255,0.08);"
                if trend_verdict.startswith("🟢"):
                    card_style = f"{base_style} border-left: 6px solid #2ecc71; background-color: rgba(46, 204, 113, 0.05);"
                else:
                    card_style = f"{base_style} border-left: 6px solid #e74c3c; background-color: rgba(231, 76, 60, 0.05);"

                html_content = f"""<div style="{card_style}">
<h4 style="margin-top: 0px; margin-bottom: 15px; font-size: 20px; letter-spacing: 0.5px;">{comm_name} <span style="font-size:14px; opacity: 0.6; font-weight: normal;">({comm_ticker})</span></h4>
<div style="display: flex; gap: 30px; margin-bottom: 20px; flex-wrap: wrap; background: rgba(128,128,128,0.1); padding: 15px; border-radius: 8px;">
<div><span style="font-size: 11px; text-transform: uppercase; letter-spacing: 1px; opacity: 0.7;">Live Price</span><br><span style="font-size: 26px; font-weight: 800; font-family: monospace;">${current_price:.2f}</span></div>
<div><span style="font-size: 11px; text-transform: uppercase; letter-spacing: 1px; opacity: 0.7;">Optimal Buy Zone</span><br><span style="font-size: 26px; font-weight: 800; color: #2ecc71; font-family: monospace;">${buy_price:.2f}</span></div>
<div><span style="font-size: 11px; text-transform: uppercase; letter-spacing: 1px; opacity: 0.7;">Optimal Sell Zone</span><br><span style="font-size: 26px; font-weight: 800; color: #e74c3c; font-family: monospace;">${sell_price:.2f}</span></div>
</div>
<h5 style="margin-top: 0; margin-bottom: 10px; font-size: 14px; text-transform: uppercase; letter-spacing: 1px;">🧠 Hermes Expert Trader AI Analysis</h5>
<ul style="margin: 0; padding-left: 20px; font-size: 14px; line-height: 1.8;">
<li><b>Market Context:</b> {bias_desc} Current live pricing is hovering at <b>${current_price:.2f}</b>.</li>
<li><b>Demand Zone (Buy):</b> A high-probability institutional order block/FVG rests at <b>${buy_price:.2f}</b> (<i>{buy_dist_pct:.1f}% below live price</i>). {action_desc if trend_verdict.startswith("🟢") else "This is the primary downside target."}</li>
<li><b>Supply Zone (Sell):</b> Unmitigated sell-side liquidity rests at <b>${sell_price:.2f}</b> (<i>{sell_dist_pct:.1f}% above live price</i>). {action_desc if not trend_verdict.startswith("🟢") else "This serves as the primary upside take-profit objective."}</li>
<li><b>Risk-to-Reward Ratio:</b> Entering a long position at the current live price to target supply, while invalidating below demand, offers an implied R:R of <b>{rr_ratio:.2f}x</b>.</li>
</ul>
</div>"""
                st.markdown(html_content, unsafe_allow_html=True)

with tab7:
    st.subheader("📈 Institutional Strategy Backtester (Rule 4)")
    st.markdown("Evaluate historical performance of the 'Rule 4' institutional setup: price closing above a key Order Block with RSI momentum support, targeting standard Risk-to-Reward multiples.")

    # Controls row
    col_bt1, col_bt2, col_bt3, col_bt4 = st.columns(4)
    with col_bt1:
        def search_asset_backtest(searchterm: str):
            return [o for o in all_ticker_options if searchterm.lower() in o.lower()] if searchterm else []
            
        bt_ticker_raw = st_searchbox(
            search_asset_backtest,
            key="backtest_searchbox",
            placeholder=f"Type search, e.g. SBIN (default: {selected_ticker})"
        )
        
        if bt_ticker_raw:
            bt_ticker = bt_ticker_raw.split(" ")[0].upper()
        else:
            bt_ticker = selected_ticker
            
        bt_ticker = bt_ticker.strip().upper()
        if "." not in bt_ticker:
            us_tickers = {"AAPL", "MSFT", "TSLA", "AMZN", "NFLX", "NVDA", "GOOG", "META", "AMD", "INTC", "QCOM", "BABA"}
            if bt_ticker not in us_tickers:
                bt_ticker = f"{bt_ticker}.NS"
                
    with col_bt2:
        bt_period = st.selectbox("Lookback Period", ["6mo", "1y", "2y"], index=1, help="Total historical data period to simulate over")
    with col_bt3:
        bt_interval = st.selectbox("Timeframe (Interval)", ["1h", "3h", "4h", "1d"], index=0, help="Candle timeframe interval")
    with col_bt4:
        bt_rr = st.slider("Risk-to-Reward Ratio (R:R)", min_value=1.0, max_value=5.0, value=2.0, step=0.1, help="Take-profit target as a multiple of risk")

    # Initialize results session state
    if "backtest_results" not in st.session_state:
        st.session_state["backtest_results"] = None

    if st.button("🚀 Run Backtest Simulation", use_container_width=True):
        with st.spinner(f"Running backtest simulation for {bt_ticker} ({bt_period}, {bt_interval}, {bt_rr} R:R)..."):
            res = run_backtest(bt_ticker, period=bt_period, interval=bt_interval, rr_ratio=bt_rr)
            st.session_state["backtest_results"] = (bt_ticker, bt_period, bt_interval, bt_rr, res)

    if st.session_state["backtest_results"] is not None:
        curr_ticker, curr_period, curr_interval, curr_rr, res = st.session_state["backtest_results"]
        
        if "error" in res:
            st.error(f"Backtest error: {res['error']}")
        else:
            st.success(f"Simulation completed for **{curr_ticker}** ({curr_period}, {curr_interval}, {curr_rr} R:R)!")
            
            # 1. Metric Cards
            st.markdown("### 📊 Performance Summary")
            m_col1, m_col2, m_col3, m_col4, m_col5 = st.columns(5)
            
            win_rate = res['win_rate']
            total_trades = res['total_trades']
            wins = res['wins']
            losses = res['losses']
            open_t = res['open']
            pf = res['profit_factor']
            dd = res['max_drawdown']
            fe = res['final_equity']
            
            # Format display strings
            pf_str = f"{pf:.2f}" if pf != float('inf') else "∞"
            
            m_col1.metric("Win Rate", f"{win_rate:.2f}%", help="Win Rate = Wins / (Wins + Losses)")
            m_col2.metric("Total Trades", f"{total_trades}", f"W: {wins} | L: {losses} | O: {open_t}", delta_color="off")
            m_col3.metric("Profit Factor", pf_str, help="Profit Factor = Gross Profits / Gross Losses")
            m_col4.metric("Max Drawdown", f"{dd:.2f}%", help="Peak-to-trough maximum paper decline")
            m_col5.metric("Compounded Return", f"{fe:.2f}", f"{(fe - 100.0):+.2f}%", help="Starting with 100.0 units of capital")
            
            # 2. Equity Curve Chart
            st.markdown("### 📈 Compounded Equity Curve")
            eq_dates = res['equity_curve']['dates']
            eq_values = res['equity_curve']['equity']
            
            # Determine color based on overall profitability
            line_color = '#26A69A' if fe >= 100.0 else '#EF5350'
            
            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=eq_dates,
                y=eq_values,
                mode='lines+markers',
                name='Compounded Equity',
                line=dict(color=line_color, width=3),
                marker=dict(size=6, symbol='circle', color=line_color),
                hovertemplate="Date: %{x}<br>Equity: %{y:.2f}<extra></extra>"
            ))
            
            # Apply same styling as main chart
            fig.update_layout(
                xaxis_rangeslider_visible=False,
                template="plotly_dark",
                height=450,
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                margin=dict(l=40, r=40, t=20, b=40),
                xaxis=dict(
                    showgrid=True, gridcolor="rgba(255,255,255,0.05)", griddash='dot',
                    showspikes=True, spikemode='across', spikethickness=1, spikedash='dot', spikecolor='#999999'
                ),
                yaxis=dict(
                    showgrid=True, gridcolor="rgba(255,255,255,0.05)", griddash='dot',
                    tickformat='.2f',
                    showspikes=True, spikemode='across', spikethickness=1, spikedash='dot', spikecolor='#999999'
                )
            )
            
            st.markdown("""
            <style>
            [data-testid="stPlotlyChart"] {
                background: linear-gradient(180deg, #1e293b 0%, #000000 100%);
                border: 2px solid #334155;
                border-radius: 12px;
                padding: 15px;
                box-shadow: 0 10px 20px rgba(0,0,0,0.2);
            }
            </style>
            """, unsafe_allow_html=True)
            
            st.plotly_chart(fig, use_container_width=True)
            
            # 3. Trade History Logs Table
            st.markdown("### 📋 Trade Logs")
            trades_list = res['trades']
            if not trades_list:
                st.info("No trades were simulated for this selection.")
            else:
                df_trades = pd.DataFrame(trades_list)
                
                # Format Dates
                df_trades['entry_date'] = pd.to_datetime(df_trades['entry_date']).dt.strftime('%Y-%m-%d %H:%M')
                df_trades['exit_date'] = pd.to_datetime(df_trades['exit_date']).dt.strftime('%Y-%m-%d %H:%M')
                
                # Round Numeric Columns
                for r_col in ['entry_price', 'SL', 'TP', 'exit_price', 'pnl', 'return_pct']:
                    if r_col in df_trades.columns:
                        df_trades[r_col] = df_trades[r_col].round(2)
                
                # Rename columns for presentation
                df_disp = df_trades.rename(columns={
                    'ticker': 'Ticker',
                    'type': 'Type',
                    'entry_date': 'Entry Date',
                    'entry_price': 'Entry Price',
                    'SL': 'Stop Loss',
                    'TP': 'Take Profit',
                    'exit_date': 'Exit Date',
                    'exit_price': 'Exit Price',
                    'pnl': 'PnL ($)',
                    'return_pct': 'Return %',
                    'outcome': 'Outcome'
                })
                
                # Color code outcome column
                def color_outcome(val):
                    if val == 'Win':
                        return 'color: #2ecc71; font-weight: bold;'
                    elif val == 'Loss':
                        return 'color: #e74c3c; font-weight: bold;'
                    return 'color: #f1c40f;'
                
                try:
                    styled_df = df_disp.style.map(color_outcome, subset=['Outcome'])
                except AttributeError:
                    styled_df = df_disp.style.applymap(color_outcome, subset=['Outcome'])
                    
                st.dataframe(styled_df, use_container_width=True)

    st.markdown("---")
    st.markdown("### 🧠 Rule 4 Strategy Execution Logics")
    
    col_logic1, col_logic2 = st.columns(2)
    with col_logic1:
        st.markdown("""
        <div style="padding: 15px; border-radius: 8px; border-left: 5px solid #3B82F6; background-color: rgba(59, 130, 246, 0.05); margin-bottom: 10px;">
            <h4 style="margin-top:0; color:#3B82F6;">📥 Entry Trigger Rules</h4>
            <ul style="margin:0; padding-left:20px; font-size:14px; line-height:1.6;">
                <li><b>Bullish (Buy) Trade Setup:</b>
                    <ul>
                        <li>Latest candle Close is higher than the top boundary of the most recent active <i>unmitigated</i> <b>Bull OB</b>.</li>
                        <li>Breakout close is fresh: within <b>5%</b> of the OB top boundary (<code>OB_Top &lt; Close &le; OB_Top * 1.05</code>).</li>
                        <li>Price is above the 50-period Volume Weighted Moving Average (<code>Close &gt; VWMA_50</code>).</li>
                        <li>14-period RSI confirms bullish momentum (<code>RSI_14 &ge; 55</code>).</li>
                    </ul>
                </li>
                <li style="margin-top:10px;"><b>Bearish (Sell) Trade Setup:</b>
                    <ul>
                        <li>Latest candle Close is lower than the bottom boundary of the most recent active <i>unmitigated</i> <b>Bear OB</b>.</li>
                        <li>Breakout close is fresh: within <b>5%</b> of the OB bottom boundary (<code>OB_Bot * 0.95 &le; Close &lt; OB_Bot</code>).</li>
                    </ul>
                </li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    with col_logic2:
        st.markdown("""
        <div style="padding: 15px; border-radius: 8px; border-left: 5px solid #10B981; background-color: rgba(16, 185, 129, 0.05); margin-bottom: 10px;">
            <h4 style="margin-top:0; color:#10B981;">📤 Exit Execution Rules</h4>
            <ul style="margin:0; padding-left:20px; font-size:14px; line-height:1.6;">
                <li><b>Long (Buy) Exit Configuration:</b>
                    <ul>
                        <li><b>Stop Loss (SL):</b> Placed exactly at the bottom boundary of the triggering Bull OB.</li>
                        <li><b>Take Profit (TP):</b> Placed dynamically based on the selected Risk-to-Reward (R:R) ratio: <code>TP = Entry + R:R * (Entry - SL)</code>.</li>
                        <li><b>Triggers:</b> Triggers if future candle's Low &le; SL (Loss) or High &ge; TP (Win).</li>
                    </ul>
                </li>
                <li style="margin-top:10px;"><b>Short (Sell) Exit Configuration:</b>
                    <ul>
                        <li><b>Stop Loss (SL):</b> Placed exactly at the top boundary of the triggering Bear OB.</li>
                        <li><b>Take Profit (TP):</b> Placed dynamically based on the selected Risk-to-Reward (R:R) ratio: <code>TP = Entry - R:R * (SL - Entry)</code>.</li>
                        <li><b>Triggers:</b> Triggers if future candle's High &ge; SL (Loss) or Low &le; TP (Win).</li>
                    </ul>
                </li>
            </ul>
        </div>
        """, unsafe_allow_html=True)

with tab8:
    st.subheader("⚡ Multi-Market Technical Scanner")
    st.markdown("**Strategy:** `1H Timeframe` | `1H RSI(9) > 1H WMA(RSI(9), 21)` | `Price > 70` | `Volume > 200k`")
    
    st.markdown("### ⚙️ Scanner Configuration")
    col_sel, col_tf, col_btn = st.columns([2, 1, 1])
    with col_sel:
        market_choice = st.selectbox(
            "Select Asset Class to Scan:",
            ["All Markets (Full Scan)", "Indian Equities (~2000)", "US Equities (Top 1000)", "Commodities"],
            key="custom_scanner_market_choice"
        )
    with col_tf:
        timeframe_choice = st.selectbox(
            "Select Timeframe:",
            ["1H", "3H", "4H", "1D"],
            key="custom_scanner_tf_choice"
        )
    with col_btn:
        st.markdown("<div style='height: 28px;'></div>", unsafe_allow_html=True)
        run_scan = st.button("🚀 Run Scanner Engine", type="primary", use_container_width=True, key="custom_scanner_run_btn")

    if run_scan:
        scan_list = []
        
        with st.spinner("Fetching latest market lists..."):
            if market_choice in ["Indian Equities (~2000)", "All Markets (Full Scan)"]:
                scan_list.extend([(t, "Indian Equities") for t in get_indian_equities()])
                
            if market_choice in ["US Equities (Top 1000)", "All Markets (Full Scan)"]:
                scan_list.extend([(t, "US Equities") for t in get_us_equities()])
                
            if market_choice in ["Commodities", "All Markets (Full Scan)"]:
                scan_list.extend([(t, "Commodities") for t in get_commodities()])

        total_assets = len(scan_list)
        
        if total_assets == 0:
            st.error("Failed to load any tickers. Please check your internet connection.")
        else:
            st.write(f"### Initializing multi-threaded scan ({timeframe_choice}) for **{total_assets}** assets...")
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            results = []
            completed = 0
            
            # Execute parallel scanning 
            with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
                future_to_ticker = {
                    executor.submit(check_stock_conditions, t[0], t[1], timeframe_choice): t 
                    for t in scan_list
                }
                
                for future in concurrent.futures.as_completed(future_to_ticker):
                    completed += 1
                    
                    # Update UI elements
                    progress_bar.progress(completed / total_assets)
                    if completed % 10 == 0 or completed == total_assets:
                        status_text.text(f"Processed {completed} of {total_assets} assets...")
                    
                    res = future.result()
                    if res:
                        results.append(res)

            st.divider()
            
            # Results rendering
            if results:
                st.success(f"Scan Complete! Found **{len(results)}** assets matching the criteria.")
                
                df_results = pd.DataFrame(results)
                
                # Display interactive dataframe
                st.dataframe(
                    df_results.style.format({
                        "Close": "{:.2f}",
                        "WMA(RSI, 21)": "{:.2f}",
                        "Volume": "{:,}",
                        "RSI (9)": "{:.2f}"
                    }),
                    use_container_width=True,
                    hide_index=True
                )
                
                # Excel Export 
                excel_data = create_formatted_excel(df_results)
                st.download_button(
                    label="📥 Download Formatted Excel Report",
                    data=excel_data,
                    file_name=f"Scanner_Results_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
                
            else:
                st.warning("Scan complete. No assets matched the strategy conditions during this execution.")

st.markdown("---")
st.markdown(
    """
    <div style='text-align: center; color: gray; padding: 10px; font-size: 14px;'>
        Created by <strong>SUDHIR THIKANE</strong> (<a href="mailto:sudhir.thikane@gmail.com" style="color: #26A69A; text-decoration: none;">sudhir.thikane@gmail.com</a>)
    </div>
    """,
    unsafe_allow_html=True
)

